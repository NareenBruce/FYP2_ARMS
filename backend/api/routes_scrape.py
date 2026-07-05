import json
import csv
import os
import tempfile

import openpyxl
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, BackgroundTasks

from core.validators import validate_gs_id, validate_university
from core.embeddings import incremental_embed_new_reviewers, init_sqlite_db, reload_experts
from scraper.scholar_scraper import scrape_single_reviewer, scrape_batch_reviewers
from models import SingleScrapeRequest, ScrapeSingleResponse, ScrapeBatchResponse

router = APIRouter(prefix="/api/scrape", tags=["Scraping"])

# Track scrape job status
scrape_status = {
    "running": False,
    "progress": "",
    "result": None
}


@router.post("/single", response_model=ScrapeSingleResponse)
async def scrape_single(req: SingleScrapeRequest):
    """Scrape a single reviewer profile and add to database."""
    from main import app_state

    # Validate inputs
    if not validate_gs_id(req.g_scholar_id):
        raise HTTPException(status_code=400, detail=f"Invalid Google Scholar ID format: '{req.g_scholar_id}'")
    if not validate_university(req.university):
        raise HTTPException(status_code=400, detail=f"Invalid university name: '{req.university}'. Must contain 'University'.")

    # Check for duplicate
    with open(app_state["db_file"], 'r', encoding='utf-8') as f:
        existing_data = json.load(f)
    existing_ids = {p['g_scholar_id'] for p in existing_data}

    if req.g_scholar_id in existing_ids:
        raise HTTPException(status_code=409, detail=f"Reviewer '{req.name}' already exists in database.")

    # Run scrape
    result = scrape_single_reviewer(req.name, req.g_scholar_id, req.university)

    if result["status"] == "verified":
        # Save to database
        with open(app_state["db_file"], 'r', encoding='utf-8') as f:
            existing_data = json.load(f)

        existing_data.append({
            "name": result["name"],
            "g_scholar_id": result["g_scholar_id"],
            "university": result["university"],
            "verified": result["verified"],
            "email": result["email"],
            "publications": result["publications"]
        })

        with open(app_state["db_file"], 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, indent=4, ensure_ascii=False)

        # Incremental embed
        incremental_embed_new_reviewers([result], app_state["model"])

        # Update SQLite
        init_sqlite_db()

        # Reload experts in app state
        app_state["experts"] = reload_experts()

        return ScrapeSingleResponse(
            status="verified",
            name=result["name"],
            g_scholar_id=result["g_scholar_id"],
            university=result["university"],
            email=result["email"],
            publications_count=len(result["publications"])
        )

    return ScrapeSingleResponse(
        status=result["status"],
        name=result.get("name", req.name),
        reason=result.get("reason", "")
    )


def _read_input_file(file_path):
    """Reads CSV or Excel file. Returns list of {name, g_scholar_id} dicts."""
    ext = os.path.splitext(file_path)[1].lower()
    entries = []

    if ext == '.csv':
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            fieldnames = [fn.strip() for fn in reader.fieldnames] if reader.fieldnames else []

            name_col = None
            id_col = None
            for fn in fieldnames:
                fn_lower = fn.lower().strip()
                if fn_lower in ('name', 'names', 'reviewer name', 'reviewer', 'scholar id', 'name'):
                    if name_col is None:
                        name_col = fn
                    else:
                        id_col = fn
                if fn_lower in ('g_scholar_id', 'gs_id', 'scholar_id', 'google scholar id', 'scholar id'):
                    id_col = fn

            if not name_col or not id_col:
                return None

            for row in reader:
                name = row.get(name_col, '').strip()
                gs_id = row.get(id_col, '').strip()
                if name and gs_id:
                    entries.append({"name": name, "g_scholar_id": gs_id})

    elif ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

        name_col = None
        id_col = None
        for idx, h in enumerate(headers):
            h_lower = h.lower().strip()
            if h_lower in ('name', 'names', 'reviewer name', 'reviewer'):
                name_col = idx
            if h_lower in ('g_scholar_id', 'gs_id', 'scholar_id', 'google scholar id', 'scholar id'):
                id_col = idx

        if name_col is None or id_col is None:
            return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_col]).strip() if row[name_col] else ''
            gs_id = str(row[id_col]).strip() if row[id_col] else ''
            if name and gs_id:
                entries.append({"name": name, "g_scholar_id": gs_id})

    return entries


@router.post("/batch", response_model=ScrapeBatchResponse)
async def scrape_batch(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    university: str = Form(...)
):
    """Upload CSV/Excel and scrape all reviewers in batch."""
    from main import app_state

    if scrape_status["running"]:
        raise HTTPException(status_code=409, detail="A scrape job is already running.")

    if not validate_university(university):
        raise HTTPException(status_code=400, detail=f"Invalid university name: '{university}'. Must contain 'University'.")

    # Save uploaded file to temp
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.csv', '.xlsx', '.xls'):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are accepted.")

    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        raw_entries = _read_input_file(tmp_path)
    finally:
        os.unlink(tmp_path)

    if raw_entries is None:
        raise HTTPException(status_code=400, detail="Could not find 'Name' and 'g_scholar_id' columns in file.")

    if not raw_entries:
        raise HTTPException(status_code=400, detail="No entries found in file.")

    # Validate and dedup
    valid_entries = [e for e in raw_entries if validate_gs_id(e['g_scholar_id'])]
    invalid_count = len(raw_entries) - len(valid_entries)

    with open(app_state["db_file"], 'r', encoding='utf-8') as f:
        existing_data = json.load(f)
    existing_ids = {p['g_scholar_id'] for p in existing_data}

    deduped_entries = [e for e in valid_entries if e['g_scholar_id'] not in existing_ids]

    if not deduped_entries:
        raise HTTPException(status_code=400, detail="No new reviewers to add (all invalid or duplicates).")

    # Run scrape (synchronous for now — could be made async with background tasks)
    results = scrape_batch_reviewers(deduped_entries, university)

    # Save verified reviewers
    if results["verified"]:
        with open(app_state["db_file"], 'r', encoding='utf-8') as f:
            existing_data = json.load(f)



        existing_data.extend(results["verified"])

        with open(app_state["db_file"], 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, indent=4, ensure_ascii=False)

        incremental_embed_new_reviewers(results["verified"], app_state["model"])
        init_sqlite_db()
        app_state["experts"] = reload_experts()

    return ScrapeBatchResponse(
        verified=len(results["verified"]),
        unverified=len(results["unverified"]),
        inactive=len(results["inactive"]),
        failed=len(results["failed"]),
        details={
            "verified": [{"name": v["name"], "g_scholar_id": v["g_scholar_id"]} for v in results["verified"]],
            "unverified": [{"name": u["name"], "reason": u["reason"]} for u in results["unverified"]],
            "inactive": [{"name": u["name"], "reason": u["reason"]} for u in results["inactive"]],
            "failed": [{"name": f["name"], "reason": f.get("reason", "Unknown")} for f in results["failed"]],
            "invalid_ids": invalid_count
        }
    )


@router.get("/status")
async def get_scrape_status():
    """Returns current scrape job status."""
    return scrape_status
